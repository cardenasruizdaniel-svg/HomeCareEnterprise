"""
HomeCare Enterprise - Importacion masiva de profesionales desde Excel
"""

import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from services import profesionales_service

COLUMNAS = [
    ("documento", "Número Documento*", 18),
    ("primer_nombre", "Primer Nombre*", 18),
    ("segundo_nombre", "Segundo Nombre", 18),
    ("primer_apellido", "Primer Apellido*", 18),
    ("segundo_apellido", "Segundo Apellido", 18),
    ("especialidad_principal", "Cargo / Especialidad*", 22),
    ("celular", "Celular", 14),
    ("telefono", "Teléfono fijo", 14),
    ("correo", "Correo electrónico", 24),
    ("direccion", "Dirección", 26),
    ("municipio", "Municipio", 18),
    ("departamento", "Departamento", 18),
    ("tipo_contrato", "Tipo de Contrato* (POR_HORAS o FIJO)", 24),
    ("valor_hora", "Valor por Hora (si es POR_HORAS)", 20),
    ("salario_fijo", "Salario Fijo Mensual (si es FIJO)", 22),
]

COLUMNAS_REQUERIDAS = ["documento", "primer_nombre", "primer_apellido", "especialidad_principal"]

AZUL = "0D6EFD"
GRIS_CLARO = "F4F6F9"


def plantilla_excel() -> bytes:

    wb = Workbook()

    hoja_instrucciones = wb.active
    hoja_instrucciones.title = "Instrucciones"

    instrucciones = [
        ("Cómo diligenciar esta plantilla", True, 14),
        ("", False, 11),
        ("1. Vaya a la pestaña 'Profesionales' (abajo).", False, 11),
        ("2. La primera fila ya tiene los nombres de columna — no la modifique ni la borre.", False, 11),
        ("3. La segunda fila es un EJEMPLO — bórrela antes de subir el archivo.", False, 11),
        ("4. Escriba un profesional por fila, empezando en la fila 2.", False, 11),
        ("5. Los campos marcados con * son obligatorios.", False, 11),
        ("6. 'Tipo de Contrato' tiene lista desplegable: POR_HORAS (se paga por hora trabajada) o FIJO (salario mensual).", False, 11),
        ("7. Si es POR_HORAS, diligencie 'Valor por Hora'. Si es FIJO, diligencie 'Salario Fijo Mensual'.", False, 11),
        ("8. Guarde el archivo (no cambie el formato .xlsx) y súbalo en el sistema.", False, 11),
    ]

    for fila, (texto, negrita, tamano) in enumerate(instrucciones, start=1):
        celda = hoja_instrucciones.cell(row=fila, column=1, value=texto)
        celda.font = Font(bold=negrita, size=tamano, color=AZUL if negrita else "000000")

    hoja_instrucciones.column_dimensions["A"].width = 100

    hoja = wb.create_sheet("Profesionales")

    for col_idx, (clave, encabezado, ancho) in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=1, column=col_idx, value=encabezado)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=AZUL)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hoja.column_dimensions[chr(64 + col_idx)].width = ancho

    hoja.row_dimensions[1].height = 32
    hoja.freeze_panes = "A2"

    ejemplo = [
        "900111222", "Laura", "", "Gómez", "", "Enfermero", "3001112233", "",
        "laura@example.com", "Cll 20", "Armenia", "Quindío", "POR_HORAS", "18000", "0",
    ]
    for col_idx, valor in enumerate(ejemplo, start=1):
        celda = hoja.cell(row=2, column=col_idx, value=valor)
        celda.font = Font(italic=True, color="888888")
        celda.fill = PatternFill("solid", fgColor=GRIS_CLARO)

    dv_contrato = DataValidation(type="list", formula1='"POR_HORAS,FIJO"', allow_blank=True)
    hoja.add_data_validation(dv_contrato)
    dv_contrato.add("M2:M1000")

    salida = io.BytesIO()
    wb.save(salida)
    return salida.getvalue()


def _leer_filas_excel(contenido_bytes: bytes):
    wb = load_workbook(io.BytesIO(contenido_bytes), data_only=True)

    hoja = wb["Profesionales"] if "Profesionales" in wb.sheetnames else wb.worksheets[0]

    filas = list(hoja.iter_rows(values_only=True))

    if not filas:
        raise ValueError("El archivo está vacío.")

    encabezados_archivo = [str(c).strip() if c else "" for c in filas[0]]
    mapa_encabezado_a_clave = {encabezado: clave for clave, encabezado, _ in COLUMNAS}
    claves_columnas = [mapa_encabezado_a_clave.get(enc, "") for enc in encabezados_archivo]

    if not any(claves_columnas):
        raise ValueError(
            "No se reconocen los encabezados del archivo. Use la plantilla descargada desde el sistema."
        )

    registros = []
    for fila in filas[1:]:
        if fila is None or all(c is None or str(c).strip() == "" for c in fila):
            continue
        registro = {}
        for clave, valor in zip(claves_columnas, fila):
            if not clave:
                continue
            registro[clave] = "" if valor is None else str(valor).strip()
        registros.append(registro)

    return registros


def importar_profesionales_excel(contenido_bytes: bytes, usuario_id=None) -> dict:

    registros = _leer_filas_excel(contenido_bytes)

    exitosos = 0
    errores = []

    for numero_fila, datos in enumerate(registros, start=2):

        try:
            for campo in COLUMNAS_REQUERIDAS:
                if not datos.get(campo):
                    raise ValueError(f"falta el campo obligatorio '{campo}'")

            datos["tipo_contrato"] = (datos.get("tipo_contrato") or "POR_HORAS").upper()
            datos["valor_hora"] = float(datos["valor_hora"]) if datos.get("valor_hora") else 0
            datos["salario_fijo"] = float(datos["salario_fijo"]) if datos.get("salario_fijo") else 0

            profesionales_service.crear(datos, usuario_id=usuario_id)
            exitosos += 1

        except Exception as error:
            errores.append({"fila": numero_fila, "documento": datos.get("documento", ""), "error": str(error)})

    return {"exitosos": exitosos, "errores": errores, "total": exitosos + len(errores)}
