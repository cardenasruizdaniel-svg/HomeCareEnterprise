"""
HomeCare Enterprise - Inventario de Insumos Médicos (módulo profesional)

Maneja el catálogo de insumos, los proveedores, los convenios
con proveedores, y los movimientos de ENTRADA (compras) y
SALIDA (insumos entregados/usados en el tratamiento de un
paciente).

Funciona como un kardex: cada movimiento queda con el SALDO
que quedó en el momento (saldo_despues), para poder reconstruir
el historial exacto de existencias en cualquier fecha, además
del stock actual.

El costo de cada insumo se calcula como COSTO PROMEDIO
PONDERADO: cada vez que entra una compra con un costo distinto
al que ya se tenía, se recalcula el promedio pesado por
cantidad -- así el valor del inventario refleja el costo real
acumulado, no solo el de la última compra.
"""

from database.database import consultar_uno
from repositories.inventario_repository import (
    ConveniosRepository,
    InsumosRepository,
    InventarioMovimientosRepository,
    ProveedoresRepository,
)

CATEGORIAS_INSUMO = [
    "Curación", "Protección personal (EPP)", "Sueros y soluciones",
    "Medicamentos", "Sondas y catéteres", "Oxigenoterapia", "Aseo y desinfección", "Otro",
]

TIPOS_CONVENIO = ["Suministro", "Servicio", "Comodato", "Exclusividad", "Otro"]


# ==========================================
# PROVEEDORES
# ==========================================

def listar_proveedores_activos():
    return [dict(p) for p in ProveedoresRepository.listar_activos()]


def listar_proveedores():
    return [dict(p) for p in ProveedoresRepository.listar_todos()]


def obtener_proveedor(proveedor_id: int):
    fila = ProveedoresRepository.obtener(proveedor_id)
    return dict(fila) if fila else None


def crear_proveedor(nombre, nit, contacto, telefono, correo, direccion) -> int:
    if not nombre:
        raise ValueError("El nombre del proveedor es obligatorio.")
    return ProveedoresRepository.crear({
        "nombre": nombre, "nit": nit or "", "contacto": contacto or "",
        "telefono": telefono or "", "correo": correo or "", "direccion": direccion or "",
    })


def desactivar_proveedor(proveedor_id: int):
    ProveedoresRepository.cambiar_estado(proveedor_id, "Inactivo")


# ==========================================
# INSUMOS (catálogo)
# ==========================================

def listar_insumos_con_stock():
    filas = [dict(i) for i in InsumosRepository.listar_con_stock()]
    for f in filas:
        f["stock_bajo"] = f["stock_actual"] < (f["stock_minimo"] or 0)
        f["stock_exceso"] = bool(f["stock_maximo"]) and f["stock_actual"] > f["stock_maximo"]
        f["valor_existencia"] = round(f["stock_actual"] * (f["costo_promedio"] or 0), 2)
    return filas


def crear_insumo(nombre, categoria, unidad_medida, stock_minimo, codigo=None, stock_maximo=0, requiere_lote_vencimiento=False) -> int:
    if not nombre:
        raise ValueError("El nombre del insumo es obligatorio.")
    return InsumosRepository.crear({
        "codigo": codigo or None, "nombre": nombre, "categoria": categoria or "Otro",
        "unidad_medida": unidad_medida or "Unidad", "stock_minimo": stock_minimo or 0,
        "stock_maximo": stock_maximo or 0, "requiere_lote_vencimiento": 1 if requiere_lote_vencimiento else 0,
    })


def desactivar_insumo(insumo_id: int):
    InsumosRepository.desactivar(insumo_id)


def actualizar_insumo(insumo_id, nombre, categoria, unidad_medida, stock_minimo, codigo=None,
                        stock_maximo=0, requiere_lote_vencimiento=False):
    if not nombre:
        raise ValueError("El nombre del insumo es obligatorio.")
    InsumosRepository.actualizar(insumo_id, {
        "codigo": codigo or None, "nombre": nombre, "categoria": categoria or "Otro",
        "unidad_medida": unidad_medida or "Unidad", "stock_minimo": stock_minimo or 0,
        "stock_maximo": stock_maximo or 0, "requiere_lote_vencimiento": 1 if requiere_lote_vencimiento else 0,
    })


CATEGORIAS_INSUMOS = CATEGORIAS_INSUMO  # alias -- se usa la misma lista de categorías que ya existía, para no tener dos listas distintas que se puedan desincronizar


def generar_plantilla_carga_masiva() -> str:
    """
    Genera el archivo Excel en blanco (con los encabezados
    correctos y una hoja de ejemplo) que se debe llenar para
    crear o actualizar varios insumos de una sola vez.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from pathlib import Path
    from core.config import EXPORTS_DIR

    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "Insumos"

    encabezados = ["codigo", "nombre", "categoria", "unidad_medida", "stock_minimo", "stock_maximo", "requiere_lote_vencimiento"]
    hoja.append(encabezados)
    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="0A8F86", end_color="0A8F86", fill_type="solid")

    hoja.append(["INS-001", "Guantes de nitrilo talla M", "Protección personal (EPP)", "Caja x100", 5, 30, "No"])
    hoja.append(["INS-002", "Gasas estériles 10x10", "Curación", "Paquete x10", 10, 50, "No"])
    hoja.append(["", "Acetaminofén 500mg", "Medicamentos", "Tableta", 20, 200, "No"])

    hoja_ayuda = libro.create_sheet("Instrucciones")
    instrucciones = [
        ["Cómo usar esta plantilla"],
        [""],
        ["- 'codigo': opcional. Si lo deja en blanco, el sistema no le asigna ninguno."],
        ["- 'nombre': obligatorio. Es la única columna que no puede quedar vacía."],
        [f"- 'categoria': una de estas: {', '.join(CATEGORIAS_INSUMOS)} (si escribe otra cosa, queda como 'Otro')."],
        ["- 'unidad_medida': texto libre, ej: Unidad, Caja x100, Frasco, Tableta, Paquete x10."],
        ["- 'stock_minimo' y 'stock_maximo': números enteros (déjelos en 0 si no aplica)."],
        ["- 'requiere_lote_vencimiento': escriba 'Si' o 'No'."],
        [""],
        ["Si el 'nombre' ya existe en el sistema (comparando sin importar mayúsculas), se ACTUALIZAN sus datos"],
        ["con lo que traiga la fila, en vez de crear un insumo duplicado."],
    ]
    for fila in instrucciones:
        hoja_ayuda.append(fila)
    hoja_ayuda["A1"].font = Font(bold=True, size=13)
    hoja_ayuda.column_dimensions["A"].width = 100

    hoja.column_dimensions["A"].width = 14
    hoja.column_dimensions["B"].width = 34
    hoja.column_dimensions["C"].width = 26
    hoja.column_dimensions["D"].width = 16
    hoja.column_dimensions["E"].width = 14
    hoja.column_dimensions["F"].width = 14
    hoja.column_dimensions["G"].width = 22

    carpeta = Path(EXPORTS_DIR) / "plantillas"
    carpeta.mkdir(parents=True, exist_ok=True)
    ruta = carpeta / "plantilla_carga_insumos.xlsx"
    libro.save(ruta)
    return str(ruta)


def cargar_insumos_desde_excel(ruta_archivo: str) -> dict:
    """
    Lee el Excel de la plantilla y crea/actualiza los insumos
    fila por fila. Si el nombre ya existe (sin importar
    mayúsculas), se actualiza ese insumo en vez de crear uno
    repetido -- así la misma plantilla también sirve para subir
    cambios de varios productos a la vez, no solo para cargar
    productos nuevos.
    """
    import openpyxl

    libro = openpyxl.load_workbook(ruta_archivo, data_only=True)
    hoja = libro["Insumos"] if "Insumos" in libro.sheetnames else libro.worksheets[0]

    filas = list(hoja.iter_rows(values_only=True))
    if not filas:
        raise ValueError("El archivo está vacío.")

    encabezados = [str(c).strip().lower() if c else "" for c in filas[0]]
    columnas_esperadas = {"nombre"}
    if not columnas_esperadas.issubset(set(encabezados)):
        raise ValueError("El archivo no tiene el formato esperado -- descargue la plantilla de nuevo y no cambie los nombres de las columnas.")

    creados, actualizados, errores = [], [], []

    for numero_fila, fila in enumerate(filas[1:], start=2):
        if not fila or all(c in (None, "") for c in fila):
            continue

        datos = dict(zip(encabezados, fila))
        nombre = str(datos.get("nombre") or "").strip()
        if not nombre:
            errores.append(f"Fila {numero_fila}: el nombre está vacío, se omitió.")
            continue

        codigo = str(datos.get("codigo") or "").strip() or None
        categoria = str(datos.get("categoria") or "").strip() or "Otro"
        if categoria not in CATEGORIAS_INSUMOS:
            categoria = "Otro"
        unidad_medida = str(datos.get("unidad_medida") or "").strip() or "Unidad"

        try:
            stock_minimo = int(datos.get("stock_minimo") or 0)
        except (TypeError, ValueError):
            stock_minimo = 0
        try:
            stock_maximo = int(datos.get("stock_maximo") or 0)
        except (TypeError, ValueError):
            stock_maximo = 0

        valor_lote = str(datos.get("requiere_lote_vencimiento") or "").strip().lower()
        requiere_lote = valor_lote in ("si", "sí", "true", "1", "x")

        existente = InsumosRepository.buscar_por_nombre_exacto(nombre)

        try:
            if existente:
                insumo_id = dict(existente)["id"]
                actualizar_insumo(insumo_id, nombre, categoria, unidad_medida, stock_minimo, codigo, stock_maximo, requiere_lote)
                actualizados.append(nombre)
            else:
                crear_insumo(nombre, categoria, unidad_medida, stock_minimo, codigo, stock_maximo, requiere_lote)
                creados.append(nombre)
        except Exception as error:
            errores.append(f"Fila {numero_fila} ({nombre}): {error}")

    return {
        "creados": creados, "actualizados": actualizados, "errores": errores,
        "total_creados": len(creados), "total_actualizados": len(actualizados), "total_errores": len(errores),
    }


def obtener_insumo(insumo_id: int):
    fila = InsumosRepository.obtener(insumo_id)
    return dict(fila) if fila else None


def stock_actual(insumo_id: int) -> int:
    return InsumosRepository.stock_actual(insumo_id)


# ==========================================
# MOVIMIENTOS (entradas / salidas) -- estilo kardex
# ==========================================

def registrar_entrada(insumo_id, cantidad, proveedor_id, numero_factura, costo_unitario, motivo, usuario_id,
                        lote=None, fecha_vencimiento=None) -> int:
    """Compra recibida de un proveedor: aumenta el stock y recalcula el costo promedio ponderado."""

    if not insumo_id or not cantidad or cantidad <= 0:
        raise ValueError("Debe indicar el insumo y una cantidad mayor a cero.")

    insumo = obtener_insumo(insumo_id)
    if not insumo:
        raise ValueError("El insumo no existe.")
    if insumo.get("requiere_lote_vencimiento") and not fecha_vencimiento:
        raise ValueError("Este insumo requiere indicar la fecha de vencimiento del lote.")

    stock_previo = stock_actual(insumo_id)
    costo_previo = insumo.get("costo_promedio") or 0
    costo_nuevo = costo_unitario or costo_previo

    # Costo promedio ponderado: (stock que ya había * su costo + lo que entra * su costo) / stock total
    if (stock_previo + cantidad) > 0:
        costo_promedio_nuevo = round(((stock_previo * costo_previo) + (cantidad * costo_nuevo)) / (stock_previo + cantidad), 4)
    else:
        costo_promedio_nuevo = costo_nuevo
    InsumosRepository.actualizar_costo_promedio(insumo_id, costo_promedio_nuevo)

    movimiento_id = InventarioMovimientosRepository.crear({
        "insumo_id": insumo_id, "tipo": "Entrada", "cantidad": cantidad,
        "proveedor_id": proveedor_id or None, "numero_factura": numero_factura or "",
        "costo_unitario": costo_nuevo, "paciente_id": None, "profesional_id": None,
        "motivo": motivo or "Compra", "lote": lote or None, "fecha_vencimiento": fecha_vencimiento or None,
        "saldo_despues": stock_previo + cantidad, "usuario_creacion": usuario_id,
    })
    return movimiento_id


def registrar_salida(insumo_id, cantidad, paciente_id, profesional_id, motivo, usuario_id) -> int:
    """Insumo entregado/usado en el tratamiento de un paciente: disminuye el stock."""

    if not insumo_id or not cantidad or cantidad <= 0:
        raise ValueError("Debe indicar el insumo y una cantidad mayor a cero.")

    stock = stock_actual(insumo_id)
    if cantidad > stock:
        raise ValueError(f"No hay suficiente stock: disponible {stock}, solicitado {cantidad}.")

    return InventarioMovimientosRepository.crear({
        "insumo_id": insumo_id, "tipo": "Salida", "cantidad": cantidad,
        "proveedor_id": None, "numero_factura": "", "costo_unitario": None,
        "paciente_id": paciente_id or None, "profesional_id": profesional_id or None,
        "motivo": motivo or "Entrega para tratamiento", "lote": None, "fecha_vencimiento": None,
        "saldo_despues": stock - cantidad, "usuario_creacion": usuario_id,
    })


def listar_movimientos_por_insumo(insumo_id: int):
    return [dict(m) for m in InventarioMovimientosRepository.listar_por_insumo(insumo_id)]


def listar_movimientos_por_paciente(paciente_id: int):
    return [dict(m) for m in InventarioMovimientosRepository.listar_por_paciente(paciente_id)]


def listar_movimientos_recientes(limite=200):
    return [dict(m) for m in InventarioMovimientosRepository.listar_todos(limite)]


# ==========================================
# INFORMES PROFESIONALES
# ==========================================

def informe_existencias():
    """Informe de existencias actuales, valorizado -- para saber qué hay y cuánto vale el inventario."""
    filas = listar_insumos_con_stock()
    total_valorizado = round(sum(f["valor_existencia"] for f in filas), 2)
    return {
        "insumos": filas,
        "total_valorizado": total_valorizado,
        "total_insumos": len(filas),
        "insumos_stock_bajo": [f for f in filas if f["stock_bajo"]],
        "insumos_stock_exceso": [f for f in filas if f["stock_exceso"]],
    }


def informe_compras(fecha_desde: str, fecha_hasta: str, proveedor_id=None):
    """Informe de compras (entradas) en un rango de fechas, con el valor total comprado."""
    filas = [dict(m) for m in InventarioMovimientosRepository.informe_compras(fecha_desde, fecha_hasta, proveedor_id)]
    for f in filas:
        f["valor_total"] = round((f["cantidad"] or 0) * (f["costo_unitario"] or 0), 2)
    total_comprado = round(sum(f["valor_total"] for f in filas), 2)

    por_proveedor = {}
    for f in filas:
        clave = f.get("proveedor") or "Sin proveedor"
        por_proveedor.setdefault(clave, 0)
        por_proveedor[clave] += f["valor_total"]

    return {
        "movimientos": filas, "total_comprado": total_comprado, "total_movimientos": len(filas),
        "por_proveedor": [{"proveedor": k, "valor": round(v, 2)} for k, v in sorted(por_proveedor.items(), key=lambda x: -x[1])],
    }


def informe_movimientos(fecha_desde: str, fecha_hasta: str, insumo_id=None, tipo=None):
    """Informe de ingresos y egresos (kardex) en un rango de fechas."""
    filas = [dict(m) for m in InventarioMovimientosRepository.informe_movimientos(fecha_desde, fecha_hasta, insumo_id, tipo)]
    total_entradas = sum(f["cantidad"] for f in filas if f["tipo"] == "Entrada")
    total_salidas = sum(f["cantidad"] for f in filas if f["tipo"] == "Salida")
    return {
        "movimientos": filas, "total_entradas": total_entradas, "total_salidas": total_salidas,
        "total_movimientos": len(filas),
    }


def alertas_vencimiento(dias=90):
    """Lotes que vencen dentro de los próximos N días -- para gestionarlos antes de que se pierdan."""
    filas = [dict(f) for f in InventarioMovimientosRepository.lotes_por_vencer(dias)]
    for f in filas:
        f["dias_restantes"] = int(f["dias_restantes"]) if f["dias_restantes"] is not None else None
        f["vencido"] = f["dias_restantes"] is not None and f["dias_restantes"] < 0
    return filas


def resumen_dashboard():
    """Resumen ejecutivo para la pantalla principal de inventario."""
    existencias = informe_existencias()
    vencimientos = alertas_vencimiento(90)
    return {
        "total_insumos": existencias["total_insumos"],
        "total_valorizado": existencias["total_valorizado"],
        "insumos_stock_bajo": len(existencias["insumos_stock_bajo"]),
        "insumos_stock_exceso": len(existencias["insumos_stock_exceso"]),
        "lotes_por_vencer": len([v for v in vencimientos if not v["vencido"]]),
        "lotes_vencidos": len([v for v in vencimientos if v["vencido"]]),
        "total_convenios_vigentes": len([c for c in listar_convenios() if c["estado"] == "Vigente"]),
    }


# ==========================================
# CONVENIOS CON PROVEEDORES
# ==========================================

def listar_convenios():
    return [dict(c) for c in ConveniosRepository.listar_todos()]


def obtener_convenio(convenio_id: int):
    fila = ConveniosRepository.obtener(convenio_id)
    return dict(fila) if fila else None


def crear_convenio(proveedor_id, numero_convenio, tipo, fecha_inicio, fecha_fin, valor, condiciones, usuario_id) -> int:
    if not proveedor_id:
        raise ValueError("Debe indicar el proveedor.")
    if tipo not in TIPOS_CONVENIO:
        raise ValueError("Tipo de convenio no válido.")
    return ConveniosRepository.crear({
        "proveedor_id": proveedor_id, "numero_convenio": numero_convenio or "",
        "tipo": tipo, "fecha_inicio": fecha_inicio or None, "fecha_fin": fecha_fin or None,
        "valor": valor or None, "condiciones": condiciones or "", "usuario_creacion": usuario_id,
    })


def finalizar_convenio(convenio_id: int):
    ConveniosRepository.cambiar_estado(convenio_id, "Finalizado")
