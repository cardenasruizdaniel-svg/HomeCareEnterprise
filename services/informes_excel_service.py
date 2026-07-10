"""
HomeCare Enterprise - Exportación de informes a Excel

Usa openpyxl para generar los archivos .xlsx descargables de
cada informe, con formato profesional (encabezado en negrita,
columnas ajustadas, congelado el encabezado para poder
desplazarse por filas largas sin perderlo de vista).
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.config import EXPORTS_DIR, RIPS_RAZON_SOCIAL

FUENTE = "Arial"

RELLENO_ENCABEZADO = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
FUENTE_ENCABEZADO = Font(name=FUENTE, bold=True, color="FFFFFF", size=10)
FUENTE_TITULO = Font(name=FUENTE, bold=True, size=14)
FUENTE_NORMAL = Font(name=FUENTE, size=10)
BORDE_FINO = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def _preparar_carpeta() -> Path:
    carpeta = Path(EXPORTS_DIR) / "informes"
    carpeta.mkdir(parents=True, exist_ok=True)
    return carpeta


def _escribir_encabezado_reporte(hoja, titulo, columnas, fila_inicio=4):
    hoja["A1"] = RIPS_RAZON_SOCIAL or "HomeCare Enterprise"
    hoja["A1"].font = FUENTE_TITULO
    hoja["A2"] = titulo
    hoja["A2"].font = Font(name=FUENTE, bold=True, size=12)
    hoja["A3"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    hoja["A3"].font = Font(name=FUENTE, italic=True, size=9, color="666666")

    for indice, columna in enumerate(columnas, start=1):
        celda = hoja.cell(row=fila_inicio, column=indice, value=columna)
        celda.font = FUENTE_ENCABEZADO
        celda.fill = RELLENO_ENCABEZADO
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = BORDE_FINO

    hoja.freeze_panes = hoja.cell(row=fila_inicio + 1, column=1)


def generar_excel_caracterizacion(pacientes: list) -> str:
    columnas = [
        ("documento", "Documento"), ("nombre_completo", "Nombre completo"),
        ("fecha_nacimiento", "Fecha nacimiento"), ("sexo", "Sexo"),
        ("eps", "EPS"), ("regimen", "Régimen"), ("tipo_cuidado", "Tipo de cuidado"),
        ("zona_ciudad", "Zona"), ("municipio", "Municipio"), ("departamento", "Departamento"),
        ("direccion", "Dirección"), ("barrio", "Barrio"), ("celular", "Celular"),
        ("diagnostico_principal", "Diagnóstico principal"), ("total_diagnosticos", "N.° diagnósticos activos"),
        ("servicios_activos", "Servicios activos"), ("profesional_principal", "Profesional principal"),
        ("fecha_ultima_nota", "Fecha última nota"), ("fecha_registro", "Fecha de registro"),
        ("estado", "Estado"),
    ]

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Caracterización"

    _escribir_encabezado_reporte(hoja, "Caracterización de Pacientes", [c[1] for c in columnas])

    fila_actual = 5
    for paciente in pacientes:
        for indice, (clave, _) in enumerate(columnas, start=1):
            celda = hoja.cell(row=fila_actual, column=indice, value=paciente.get(clave) or "")
            celda.font = FUENTE_NORMAL
            celda.border = BORDE_FINO
        fila_actual += 1

    for indice, (clave, nombre) in enumerate(columnas, start=1):
        letra = get_column_letter(indice)
        hoja.column_dimensions[letra].width = max(len(nombre) + 2, 16)

    carpeta = _preparar_carpeta()
    marca_tiempo = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = carpeta / f"caracterizacion_pacientes_{marca_tiempo}.xlsx"
    libro.save(ruta)
    return str(ruta)


def generar_excel_resumen_zonas(por_zona: list, por_municipio: list, por_eps: list) -> str:
    libro = Workbook()

    hoja_zona = libro.active
    hoja_zona.title = "Por Zona"
    _escribir_encabezado_reporte(hoja_zona, "Pacientes por Zona de la Ciudad", ["Zona", "Total pacientes", "Ventilados"])
    fila = 5
    for z in por_zona:
        hoja_zona.cell(row=fila, column=1, value=z["zona"]).font = FUENTE_NORMAL
        hoja_zona.cell(row=fila, column=2, value=z["total_pacientes"]).font = FUENTE_NORMAL
        hoja_zona.cell(row=fila, column=3, value=z["ventilados"]).font = FUENTE_NORMAL
        fila += 1
    hoja_zona.column_dimensions["A"].width = 25
    hoja_zona.column_dimensions["B"].width = 18
    hoja_zona.column_dimensions["C"].width = 14

    hoja_municipio = libro.create_sheet("Por Municipio")
    _escribir_encabezado_reporte(hoja_municipio, "Pacientes por Municipio", ["Municipio", "Total pacientes"])
    fila = 5
    for m in por_municipio:
        hoja_municipio.cell(row=fila, column=1, value=m["municipio"]).font = FUENTE_NORMAL
        hoja_municipio.cell(row=fila, column=2, value=m["total_pacientes"]).font = FUENTE_NORMAL
        fila += 1
    hoja_municipio.column_dimensions["A"].width = 25
    hoja_municipio.column_dimensions["B"].width = 18

    hoja_eps = libro.create_sheet("Por EPS")
    _escribir_encabezado_reporte(hoja_eps, "Pacientes por EPS", ["EPS", "Total pacientes"])
    fila = 5
    for e in por_eps:
        hoja_eps.cell(row=fila, column=1, value=e["eps"]).font = FUENTE_NORMAL
        hoja_eps.cell(row=fila, column=2, value=e["total_pacientes"]).font = FUENTE_NORMAL
        fila += 1
    hoja_eps.column_dimensions["A"].width = 25
    hoja_eps.column_dimensions["B"].width = 18

    carpeta = _preparar_carpeta()
    marca_tiempo = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = carpeta / f"resumen_zonas_{marca_tiempo}.xlsx"
    libro.save(ruta)
    return str(ruta)


def generar_excel_equipo_profesional(profesionales: list) -> str:
    columnas = [
        ("documento", "Documento"), ("nombre_completo", "Nombre completo"),
        ("especialidad_principal", "Especialidad"), ("tipo_vinculacion", "Vinculación"),
        ("estado", "Estado"), ("celular", "Celular"), ("correo", "Correo"),
        ("pacientes_asignados", "Pacientes asignados"),
    ]

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Equipo"
    _escribir_encabezado_reporte(hoja, "Equipo Profesional", [c[1] for c in columnas])

    fila_actual = 5
    for profesional in profesionales:
        for indice, (clave, _) in enumerate(columnas, start=1):
            celda = hoja.cell(row=fila_actual, column=indice, value=profesional.get(clave) or "")
            celda.font = FUENTE_NORMAL
            celda.border = BORDE_FINO
        fila_actual += 1

    for indice, (clave, nombre) in enumerate(columnas, start=1):
        letra = get_column_letter(indice)
        hoja.column_dimensions[letra].width = max(len(nombre) + 2, 16)

    carpeta = _preparar_carpeta()
    marca_tiempo = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = carpeta / f"equipo_profesional_{marca_tiempo}.xlsx"
    libro.save(ruta)
    return str(ruta)
