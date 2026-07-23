"""
HomeCare Enterprise - Inventario de Insumos Médicos (módulo profesional)

Maneja el catálogo de insumos, los proveedores, los convenios
con proveedores, y los movimientos de ENTRADA (compras) y
SALIDA (insumos entregados/usados en el tratamiento de un
paciente).

Funciona como un kardex: cada movimiento queda con el SALDO
que quedó en el momento (saldo_despues), para poder reconstruir
el historial exacto de existencias en cualquier fecha, además
del stock actual.

El costo de cada insumo se calcula como COSTO PROMEDIO
PONDERADO: cada vez que entra una compra con un costo distinto
al que ya se tenía, se recalcula el promedio pesado por
cantidad -- así el valor del inventario refleja el costo real
acumulado, no solo el de la última compra.
"""

from database.database import consultar_uno
from repositories.inventario_repository import (
    ConveniosRepository,
    InsumosRepository,
    InventarioMovimientosRepository,
    ProveedoresRepository,
)

CATEGORIAS_INSUMO = [
    "Curación", "Protección personal (EPP)", "Sueros y soluciones",
    "Medicamentos", "Sondas y catéteres", "Oxigenoterapia", "Aseo y desinfección", "Otro",
]

TIPOS_CONVENIO = ["Suministro", "Servicio", "Comodato", "Exclusividad", "Otro"]

# Condición de venta -- clasificación oficial que usa el INVIMA en
# Colombia para los medicamentos. Los de "Control especial" son los
# que además hay que reportar al Fondo Nacional de Estupefacientes
# (opioides, benzodiacepinas, y similares) -- estos requieren un
# control más estricto de trazabilidad.
CONDICIONES_VENTA = [
    "Venta libre",
    "Venta con fórmula médica",
    "Control especial (Fondo Nacional de Estupefacientes)",
    "Vigilada",
]


# ==========================================
# PROVEEDORES
# ==========================================

def listar_proveedores_activos():
    return [dict(p) for p in ProveedoresRepository.listar_activos()]


def listar_proveedores():
    return [dict(p) for p in ProveedoresRepository.listar_todos()]


def obtener_proveedor(proveedor_id: int):
    fila = ProveedoresRepository.obtener(proveedor_id)
    return dict(fila) if fila else None


def crear_proveedor(nombre, nit, contacto, telefono, correo, direccion) -> int:
    if not nombre:
        raise ValueError("El nombre del proveedor es obligatorio.")
    return ProveedoresRepository.crear({
        "nombre": nombre, "nit": nit or "", "contacto": contacto or "",
        "telefono": telefono or "", "correo": correo or "", "direccion": direccion or "",
    })


def desactivar_proveedor(proveedor_id: int):
    ProveedoresRepository.cambiar_estado(proveedor_id, "Inactivo")


# ==========================================
# INSUMOS (catálogo)
# ==========================================

def listar_insumos_con_stock():
    filas = [dict(i) for i in InsumosRepository.listar_con_stock()]
    for f in filas:
        f["stock_bajo"] = f["stock_actual"] < (f["stock_minimo"] or 0)
        f["stock_exceso"] = bool(f["stock_maximo"]) and f["stock_actual"] > f["stock_maximo"]
        f["valor_existencia"] = round(f["stock_actual"] * (f["costo_promedio"] or 0), 2)
    return filas


def crear_insumo(nombre, categoria, unidad_medida, stock_minimo, codigo=None, stock_maximo=0,
                   requiere_lote_vencimiento=False, codigo_barras=None,
                   registro_invima=None, titular_registro_sanitario=None, principio_activo=None,
                   concentracion=None, forma_farmaceutica=None, condicion_venta=None, requiere_cadena_frio=False) -> int:
    if not nombre:
        raise ValueError("El nombre del insumo es obligatorio.")
    if categoria == "Medicamentos" and condicion_venta and "Control especial" in condicion_venta and not registro_invima:
        raise ValueError("Los medicamentos de control especial deben tener registrado su número de registro sanitario INVIMA.")
    return InsumosRepository.crear({
        "codigo": codigo or None, "codigo_barras": codigo_barras or None, "nombre": nombre, "categoria": categoria or "Otro",
        "unidad_medida": unidad_medida or "Unidad", "stock_minimo": stock_minimo or 0,
        "stock_maximo": stock_maximo or 0, "requiere_lote_vencimiento": 1 if requiere_lote_vencimiento else 0,
        "registro_invima": registro_invima or None, "titular_registro_sanitario": titular_registro_sanitario or None,
        "principio_activo": principio_activo or None, "concentracion": concentracion or None,
        "forma_farmaceutica": forma_farmaceutica or None, "condicion_venta": condicion_venta or None,
        "requiere_cadena_frio": 1 if requiere_cadena_frio else 0,
    })


def desactivar_insumo(insumo_id: int):
    InsumosRepository.desactivar(insumo_id)


def actualizar_insumo(insumo_id, nombre, categoria, unidad_medida, stock_minimo, codigo=None,
                        stock_maximo=0, requiere_lote_vencimiento=False, codigo_barras=None,
                        registro_invima=None, titular_registro_sanitario=None, principio_activo=None,
                        concentracion=None, forma_farmaceutica=None, condicion_venta=None, requiere_cadena_frio=False):
    if not nombre:
        raise ValueError("El nombre del insumo es obligatorio.")
    if categoria == "Medicamentos" and condicion_venta and "Control especial" in condicion_venta and not registro_invima:
        raise ValueError("Los medicamentos de control especial deben tener registrado su número de registro sanitario INVIMA.")
    InsumosRepository.actualizar(insumo_id, {
        "codigo": codigo or None, "codigo_barras": codigo_barras or None, "nombre": nombre, "categoria": categoria or "Otro",
        "unidad_medida": unidad_medida or "Unidad", "stock_minimo": stock_minimo or 0,
        "stock_maximo": stock_maximo or 0, "requiere_lote_vencimiento": 1 if requiere_lote_vencimiento else 0,
        "registro_invima": registro_invima or None, "titular_registro_sanitario": titular_registro_sanitario or None,
        "principio_activo": principio_activo or None, "concentracion": concentracion or None,
        "forma_farmaceutica": forma_farmaceutica or None, "condicion_venta": condicion_venta or None,
        "requiere_cadena_frio": 1 if requiere_cadena_frio else 0,
    })


def buscar_insumo_por_codigo(texto: str):
    """Busca un insumo por su código interno O por su código de barras -- para cuando se escanea o se digita cualquiera de los dos."""
    from database.database import consultar_uno
    if not texto:
        return None
    fila = consultar_uno(
        "SELECT * FROM insumos WHERE activo=1 AND (codigo=? OR codigo_barras=?) LIMIT 1",
        (texto.strip(), texto.strip()),
    )
    return dict(fila) if fila else None


def buscar_productos(texto: str, limite: int = 15):
    """
    Búsqueda general de productos -- por código interno, código
    de barras, o nombre (parcial) -- para usar en cualquier
    pantalla donde haga falta encontrar un insumo rápido: el
    conteo físico (escaneando o escribiendo), y también al
    registrar una entrada o una salida. Si el texto coincide
    EXACTO con un código o código de barras, ese resultado va
    de primero.
    """
    from database.database import consultar_todos
    if not texto or len(texto.strip()) < 2:
        return []
    texto = texto.strip()
    filas = consultar_todos(
        """
        SELECT *,
            CASE WHEN codigo = ? OR codigo_barras = ? THEN 0 ELSE 1 END AS orden_coincidencia
        FROM insumos
        WHERE activo=1 AND (
            codigo LIKE ? OR codigo_barras LIKE ? OR nombre LIKE ?
        )
        ORDER BY orden_coincidencia, nombre
        LIMIT ?
        """,
        (texto, texto, f"%{texto}%", f"%{texto}%", f"%{texto}%", limite),
    )
    resultado = [dict(f) for f in filas]
    for r in resultado:
        r["stock_actual"] = stock_actual(r["id"])
    return resultado


CATEGORIAS_INSUMOS = CATEGORIAS_INSUMO  # alias -- se usa la misma lista de categorías que ya existía, para no tener dos listas distintas que se puedan desincronizar


def generar_plantilla_carga_masiva() -> str:
    """
    Genera el archivo Excel en blanco (con los encabezados
    correctos y una hoja de ejemplo) que se debe llenar para
    crear o actualizar varios insumos de una sola vez.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from pathlib import Path
    from core.config import EXPORTS_DIR

    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "Insumos"

    encabezados = ["codigo", "nombre", "categoria", "unidad_medida", "stock_minimo", "stock_maximo", "requiere_lote_vencimiento"]
    hoja.append(encabezados)
    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="0A8F86", end_color="0A8F86", fill_type="solid")

    hoja.append(["INS-001", "Guantes de nitrilo talla M", "Protección personal (EPP)", "Caja x100", 5, 30, "No"])
    hoja.append(["INS-002", "Gasas estériles 10x10", "Curación", "Paquete x10", 10, 50, "No"])
    hoja.append(["", "Acetaminofén 500mg", "Medicamentos", "Tableta", 20, 200, "No"])

    hoja_ayuda = libro.create_sheet("Instrucciones")
    instrucciones = [
        ["Cómo usar esta plantilla"],
        [""],
        ["- 'codigo': opcional. Si lo deja en blanco, el sistema no le asigna ninguno."],
        ["- 'nombre': obligatorio. Es la única columna que no puede quedar vacía."],
        [f"- 'categoria': una de estas: {', '.join(CATEGORIAS_INSUMOS)} (si escribe otra cosa, queda como 'Otro')."],
        ["- 'unidad_medida': texto libre, ej: Unidad, Caja x100, Frasco, Tableta, Paquete x10."],
        ["- 'stock_minimo' y 'stock_maximo': números enteros (déjelos en 0 si no aplica)."],
        ["- 'requiere_lote_vencimiento': escriba 'Si' o 'No'."],
        [""],
        ["Si el 'nombre' ya existe en el sistema (comparando sin importar mayúsculas), se ACTUALIZAN sus datos"],
        ["con lo que traiga la fila, en vez de crear un insumo duplicado."],
    ]
    for fila in instrucciones:
        hoja_ayuda.append(fila)
    hoja_ayuda["A1"].font = Font(bold=True, size=13)
    hoja_ayuda.column_dimensions["A"].width = 100

    hoja.column_dimensions["A"].width = 14
    hoja.column_dimensions["B"].width = 34
    hoja.column_dimensions["C"].width = 26
    hoja.column_dimensions["D"].width = 16
    hoja.column_dimensions["E"].width = 14
    hoja.column_dimensions["F"].width = 14
    hoja.column_dimensions["G"].width = 22

    carpeta = Path(EXPORTS_DIR) / "plantillas"
    carpeta.mkdir(parents=True, exist_ok=True)
    ruta = carpeta / "plantilla_carga_insumos.xlsx"
    libro.save(ruta)
    return str(ruta)


def cargar_insumos_desde_excel(ruta_archivo: str) -> dict:
    """
    Lee el Excel de la plantilla y crea/actualiza los insumos
    fila por fila. Si el nombre ya existe (sin importar
    mayúsculas), se actualiza ese insumo en vez de crear uno
    repetido -- así la misma plantilla también sirve para subir
    cambios de varios productos a la vez, no solo para cargar
    productos nuevos.
    """
    import openpyxl

    libro = openpyxl.load_workbook(ruta_archivo, data_only=True)
    hoja = libro["Insumos"] if "Insumos" in libro.sheetnames else libro.worksheets[0]

    filas = list(hoja.iter_rows(values_only=True))
    if not filas:
        raise ValueError("El archivo está vacío.")

    encabezados = [str(c).strip().lower() if c else "" for c in filas[0]]
    columnas_esperadas = {"nombre"}
    if not columnas_esperadas.issubset(set(encabezados)):
        raise ValueError("El archivo no tiene el formato esperado -- descargue la plantilla de nuevo y no cambie los nombres de las columnas.")

    creados, actualizados, errores = [], [], []

    for numero_fila, fila in enumerate(filas[1:], start=2):
        if not fila or all(c in (None, "") for c in fila):
            continue

        datos = dict(zip(encabezados, fila))
        nombre = str(datos.get("nombre") or "").strip()
        if not nombre:
            errores.append(f"Fila {numero_fila}: el nombre está vacío, se omitió.")
            continue

        codigo = str(datos.get("codigo") or "").strip() or None
        categoria = str(datos.get("categoria") or "").strip() or "Otro"
        if categoria not in CATEGORIAS_INSUMOS:
            categoria = "Otro"
        unidad_medida = str(datos.get("unidad_medida") or "").strip() or "Unidad"

        try:
            stock_minimo = int(datos.get("stock_minimo") or 0)
        except (TypeError, ValueError):
            stock_minimo = 0
        try:
            stock_maximo = int(datos.get("stock_maximo") or 0)
        except (TypeError, ValueError):
            stock_maximo = 0

        valor_lote = str(datos.get("requiere_lote_vencimiento") or "").strip().lower()
        requiere_lote = valor_lote in ("si", "sí", "true", "1", "x")

        existente = InsumosRepository.buscar_por_nombre_exacto(nombre)

        try:
            if existente:
                insumo_id = dict(existente)["id"]
                actualizar_insumo(insumo_id, nombre, categoria, unidad_medida, stock_minimo, codigo, stock_maximo, requiere_lote)
                actualizados.append(nombre)
            else:
                crear_insumo(nombre, categoria, unidad_medida, stock_minimo, codigo, stock_maximo, requiere_lote)
                creados.append(nombre)
        except Exception as error:
            errores.append(f"Fila {numero_fila} ({nombre}): {error}")

    return {
        "creados": creados, "actualizados": actualizados, "errores": errores,
        "total_creados": len(creados), "total_actualizados": len(actualizados), "total_errores": len(errores),
    }


def obtener_insumo(insumo_id: int):
    fila = InsumosRepository.obtener(insumo_id)
    return dict(fila) if fila else None


def stock_actual(insumo_id: int) -> int:
    return InsumosRepository.stock_actual(insumo_id)


# ==========================================
# MOVIMIENTOS (entradas / salidas) -- estilo kardex
# ==========================================

def registrar_entrada(insumo_id, cantidad, proveedor_id, numero_factura, costo_unitario, motivo, usuario_id,
                        lote=None, fecha_vencimiento=None) -> int:
    """Compra recibida de un proveedor: aumenta el stock y recalcula el costo promedio ponderado."""

    if not insumo_id or not cantidad or cantidad <= 0:
        raise ValueError("Debe indicar el insumo y una cantidad mayor a cero.")

    insumo = obtener_insumo(insumo_id)
    if not insumo:
        raise ValueError("El insumo no existe.")
    if insumo.get("requiere_lote_vencimiento") and not fecha_vencimiento:
        raise ValueError("Este insumo requiere indicar la fecha de vencimiento del lote.")

    stock_previo = stock_actual(insumo_id)
    costo_previo = insumo.get("costo_promedio") or 0
    costo_nuevo = costo_unitario or costo_previo

    # Costo promedio ponderado: (stock que ya había * su costo + lo que entra * su costo) / stock total
    if (stock_previo + cantidad) > 0:
        costo_promedio_nuevo = round(((stock_previo * costo_previo) + (cantidad * costo_nuevo)) / (stock_previo + cantidad), 4)
    else:
        costo_promedio_nuevo = costo_nuevo
    InsumosRepository.actualizar_costo_promedio(insumo_id, costo_promedio_nuevo)

    movimiento_id = InventarioMovimientosRepository.crear({
        "insumo_id": insumo_id, "tipo": "Entrada", "cantidad": cantidad,
        "proveedor_id": proveedor_id or None, "numero_factura": numero_factura or "",
        "costo_unitario": costo_nuevo, "paciente_id": None, "profesional_id": None,
        "motivo": motivo or "Compra", "lote": lote or None, "fecha_vencimiento": fecha_vencimiento or None,
        "saldo_despues": stock_previo + cantidad, "usuario_creacion": usuario_id,
    })
    return movimiento_id


def registrar_salida(insumo_id, cantidad, paciente_id, profesional_id, motivo, usuario_id) -> int:
    """Insumo entregado/usado en el tratamiento de un paciente: disminuye el stock."""

    if not insumo_id or not cantidad or cantidad <= 0:
        raise ValueError("Debe indicar el insumo y una cantidad mayor a cero.")

    stock = stock_actual(insumo_id)
    if cantidad > stock:
        raise ValueError(f"No hay suficiente stock: disponible {stock}, solicitado {cantidad}.")

    return InventarioMovimientosRepository.crear({
        "insumo_id": insumo_id, "tipo": "Salida", "cantidad": cantidad,
        "proveedor_id": None, "numero_factura": "", "costo_unitario": None,
        "paciente_id": paciente_id or None, "profesional_id": profesional_id or None,
        "motivo": motivo or "Entrega para tratamiento", "lote": None, "fecha_vencimiento": None,
        "saldo_despues": stock - cantidad, "usuario_creacion": usuario_id,
    })


def listar_movimientos_por_insumo(insumo_id: int):
    return [dict(m) for m in InventarioMovimientosRepository.listar_por_insumo(insumo_id)]


# ==========================================
# CONTEO FÍSICO DE INVENTARIO
# ==========================================

def iniciar_conteo_fisico(usuario_id, observaciones="") -> int:
    """
    Abre un nuevo conteo físico: toma una "foto" de las
    existencias que el sistema cree tener en este momento, para
    cada insumo activo -- esa es la base contra la que se va a
    comparar lo que se cuente de verdad en bodega.
    """
    from database.database import ejecutar, consultar_uno

    conteo_id = ejecutar(
        "INSERT INTO conteos_fisicos_inventario(observaciones, usuario_inicio) VALUES (?, ?)",
        (observaciones or "", usuario_id),
    )

    insumos = listar_insumos_con_stock()
    for insumo in insumos:
        ejecutar(
            "INSERT INTO conteos_fisicos_detalle(conteo_id, insumo_id, cantidad_sistema) VALUES (?, ?, ?)",
            (conteo_id, insumo["id"], insumo["stock_actual"]),
        )

    return conteo_id


def listar_conteos_fisicos():
    from database.database import consultar_todos
    filas = consultar_todos("SELECT * FROM conteos_fisicos_inventario ORDER BY fecha_inicio DESC")
    return [dict(f) for f in filas]


def obtener_conteo_fisico(conteo_id: int):
    from database.database import consultar_uno
    fila = consultar_uno("SELECT * FROM conteos_fisicos_inventario WHERE id=?", (conteo_id,))
    return dict(fila) if fila else None


def listar_detalle_conteo(conteo_id: int):
    from database.database import consultar_todos
    filas = consultar_todos(
        """
        SELECT d.*, i.nombre, i.codigo, i.codigo_barras, i.unidad_medida, i.categoria
        FROM conteos_fisicos_detalle d
        JOIN insumos i ON i.id = d.insumo_id
        WHERE d.conteo_id=?
        ORDER BY i.nombre
        """,
        (conteo_id,),
    )
    return [dict(f) for f in filas]


def generar_plantilla_conteo(conteo_id: int) -> str:
    """Excel para imprimir o llenar en bodega: cada insumo con su existencia actual del sistema, y una columna vacía para anotar el conteo físico."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from pathlib import Path
    from core.config import EXPORTS_DIR

    detalle = listar_detalle_conteo(conteo_id)

    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "Conteo"

    encabezados = ["detalle_id", "codigo", "codigo_barras", "nombre", "categoria", "unidad_medida", "cantidad_sistema", "cantidad_fisica"]
    hoja.append(encabezados)
    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="0A8F86", end_color="0A8F86", fill_type="solid")

    for item in detalle:
        hoja.append([
            item["id"], item.get("codigo") or "", item.get("codigo_barras") or "", item["nombre"],
            item.get("categoria") or "", item.get("unidad_medida") or "", item["cantidad_sistema"], None,
        ])

    hoja.column_dimensions["A"].width = 12
    hoja.column_dimensions["B"].width = 14
    hoja.column_dimensions["C"].width = 16
    hoja.column_dimensions["D"].width = 36
    hoja.column_dimensions["E"].width = 22
    hoja.column_dimensions["F"].width = 14
    hoja.column_dimensions["G"].width = 16
    hoja.column_dimensions["H"].width = 16

    hoja_ayuda = libro.create_sheet("Instrucciones")
    hoja_ayuda.append(["Cómo usar esta plantilla"])
    hoja_ayuda.append([""])
    hoja_ayuda.append(["- NO modifique ni borre la columna 'detalle_id' -- es la que enlaza cada fila con este conteo."])
    hoja_ayuda.append(["- Cuente físicamente cada insumo en bodega y escriba la cantidad real en la columna 'cantidad_fisica'."])
    hoja_ayuda.append(["- Deje en blanco 'cantidad_fisica' en los insumos que no alcance a contar -- se pueden completar después."])
    hoja_ayuda.append(["- Al terminar, suba este mismo archivo en la pantalla del conteo para registrar los resultados."])
    hoja_ayuda["A1"].font = Font(bold=True, size=13)
    hoja_ayuda.column_dimensions["A"].width = 100

    carpeta = Path(EXPORTS_DIR) / "conteos_fisicos"
    carpeta.mkdir(parents=True, exist_ok=True)
    ruta = carpeta / f"plantilla_conteo_{conteo_id}.xlsx"
    libro.save(ruta)
    return str(ruta)


def cargar_conteo_desde_excel(conteo_id: int, ruta_archivo: str) -> dict:
    """Lee la plantilla ya diligenciada y registra la cantidad física contada de cada insumo."""
    import openpyxl
    from database.database import ejecutar

    libro = openpyxl.load_workbook(ruta_archivo, data_only=True)
    hoja = libro["Conteo"] if "Conteo" in libro.sheetnames else libro.worksheets[0]

    filas = list(hoja.iter_rows(values_only=True))
    if not filas:
        raise ValueError("El archivo está vacío.")
    encabezados = [str(c).strip().lower() if c else "" for c in filas[0]]
    if "detalle_id" not in encabezados or "cantidad_fisica" not in encabezados:
        raise ValueError("El archivo no tiene el formato esperado -- descargue la plantilla de nuevo y no cambie los nombres de las columnas.")

    indice_id = encabezados.index("detalle_id")
    indice_cantidad = encabezados.index("cantidad_fisica")

    actualizados, errores = 0, []
    for numero_fila, fila in enumerate(filas[1:], start=2):
        if not fila or fila[indice_id] is None:
            continue
        detalle_id = fila[indice_id]
        cantidad_fisica = fila[indice_cantidad]
        if cantidad_fisica is None or cantidad_fisica == "":
            continue
        try:
            cantidad_fisica = int(cantidad_fisica)
        except (TypeError, ValueError):
            errores.append(f"Fila {numero_fila}: cantidad física no es un número válido.")
            continue

        registrar_conteo_item(int(detalle_id), cantidad_fisica, solo_del_conteo=conteo_id)
        actualizados += 1

    return {"actualizados": actualizados, "errores": errores}


def registrar_conteo_item(detalle_id: int, cantidad_fisica: int, solo_del_conteo=None):
    """Registra (o corrige) la cantidad física contada de UN insumo dentro de un conteo -- para diligenciarlo directamente desde la pantalla, sin Excel."""
    from database.database import ejecutar, consultar_uno

    fila = consultar_uno("SELECT * FROM conteos_fisicos_detalle WHERE id=?", (detalle_id,))
    if not fila:
        raise ValueError("El detalle de conteo indicado no existe.")
    fila = dict(fila)
    if solo_del_conteo and fila["conteo_id"] != solo_del_conteo:
        raise ValueError("Ese detalle no pertenece a este conteo.")

    diferencia = int(cantidad_fisica) - fila["cantidad_sistema"]
    ejecutar(
        "UPDATE conteos_fisicos_detalle SET cantidad_fisica=?, diferencia=? WHERE id=?",
        (int(cantidad_fisica), diferencia, detalle_id),
    )
    return {"detalle_id": detalle_id, "cantidad_sistema": fila["cantidad_sistema"], "cantidad_fisica": int(cantidad_fisica), "diferencia": diferencia}


def registrar_conteo_por_insumo(conteo_id: int, insumo_id: int, cantidad_fisica: int):
    """
    Igual que 'registrar_conteo_item', pero identificando el
    insumo directamente (no la fila de detalle) -- es lo que usa
    la pantalla de escaneo: se busca el producto por su código
    de barras/código/nombre, y con su id se registra el conteo
    de una vez, sin que la persona tenga que saber a qué fila de
    detalle corresponde.
    """
    from database.database import consultar_uno

    fila = consultar_uno(
        "SELECT id FROM conteos_fisicos_detalle WHERE conteo_id=? AND insumo_id=?",
        (conteo_id, insumo_id),
    )
    if not fila:
        raise ValueError("Este producto no hace parte de este conteo (puede que se haya creado después de iniciarlo).")

    return registrar_conteo_item(dict(fila)["id"], cantidad_fisica, solo_del_conteo=conteo_id)


def aplicar_ajustes_conteo(conteo_id: int, usuario_id) -> dict:
    """
    Cierra el conteo: para cada insumo donde el conteo físico
    fue distinto al del sistema, genera el movimiento de ajuste
    correspondiente (entrada si sobró, salida si faltó), y deja
    el conteo marcado como Cerrado. Los insumos que no se
    alcanzaron a contar (cantidad_fisica vacía) simplemente no
    se tocan.
    """
    from database.database import ejecutar

    detalle = listar_detalle_conteo(conteo_id)
    ajustes_aplicados = []

    for item in detalle:
        if item["cantidad_fisica"] is None or item["ajustado"]:
            continue
        diferencia = item["diferencia"]
        if diferencia == 0:
            ejecutar("UPDATE conteos_fisicos_detalle SET ajustado=1 WHERE id=?", (item["id"],))
            continue

        if diferencia > 0:
            InventarioMovimientosRepository.crear({
                "insumo_id": item["insumo_id"], "tipo": "Entrada", "cantidad": diferencia,
                "proveedor_id": None, "numero_factura": "", "costo_unitario": None,
                "paciente_id": None, "profesional_id": None,
                "motivo": f"Ajuste por conteo físico #{conteo_id} (sobrante)", "lote": None, "fecha_vencimiento": None,
                "saldo_despues": item["cantidad_fisica"], "usuario_creacion": usuario_id,
            })
        else:
            InventarioMovimientosRepository.crear({
                "insumo_id": item["insumo_id"], "tipo": "Salida", "cantidad": abs(diferencia),
                "proveedor_id": None, "numero_factura": "", "costo_unitario": None,
                "paciente_id": None, "profesional_id": None,
                "motivo": f"Ajuste por conteo físico #{conteo_id} (faltante)", "lote": None, "fecha_vencimiento": None,
                "saldo_despues": item["cantidad_fisica"], "usuario_creacion": usuario_id,
            })

        ejecutar("UPDATE conteos_fisicos_detalle SET ajustado=1 WHERE id=?", (item["id"],))
        ajustes_aplicados.append({"insumo": item["nombre"], "diferencia": diferencia})

    ejecutar(
        "UPDATE conteos_fisicos_inventario SET estado='Cerrado', fecha_cierre=CURRENT_TIMESTAMP, usuario_cierre=? WHERE id=?",
        (usuario_id, conteo_id),
    )

    return {"ajustes_aplicados": ajustes_aplicados, "total_ajustes": len(ajustes_aplicados)}


def listar_movimientos_por_paciente(paciente_id: int):
    return [dict(m) for m in InventarioMovimientosRepository.listar_por_paciente(paciente_id)]


def listar_movimientos_recientes(limite=200):
    return [dict(m) for m in InventarioMovimientosRepository.listar_todos(limite)]


# ==========================================
# INFORMES PROFESIONALES
# ==========================================

def informe_existencias():
    """Informe de existencias actuales, valorizado -- para saber qué hay y cuánto vale el inventario."""
    filas = listar_insumos_con_stock()
    total_valorizado = round(sum(f["valor_existencia"] for f in filas), 2)
    return {
        "insumos": filas,
        "total_valorizado": total_valorizado,
        "total_insumos": len(filas),
        "insumos_stock_bajo": [f for f in filas if f["stock_bajo"]],
        "insumos_stock_exceso": [f for f in filas if f["stock_exceso"]],
    }


def informe_compras(fecha_desde: str, fecha_hasta: str, proveedor_id=None):
    """Informe de compras (entradas) en un rango de fechas, con el valor total comprado."""
    filas = [dict(m) for m in InventarioMovimientosRepository.informe_compras(fecha_desde, fecha_hasta, proveedor_id)]
    for f in filas:
        f["valor_total"] = round((f["cantidad"] or 0) * (f["costo_unitario"] or 0), 2)
    total_comprado = round(sum(f["valor_total"] for f in filas), 2)

    por_proveedor = {}
    for f in filas:
        clave = f.get("proveedor") or "Sin proveedor"
        por_proveedor.setdefault(clave, 0)
        por_proveedor[clave] += f["valor_total"]

    return {
        "movimientos": filas, "total_comprado": total_comprado, "total_movimientos": len(filas),
        "por_proveedor": [{"proveedor": k, "valor": round(v, 2)} for k, v in sorted(por_proveedor.items(), key=lambda x: -x[1])],
    }


def informe_movimientos(fecha_desde: str, fecha_hasta: str, insumo_id=None, tipo=None):
    """Informe de ingresos y egresos (kardex) en un rango de fechas."""
    filas = [dict(m) for m in InventarioMovimientosRepository.informe_movimientos(fecha_desde, fecha_hasta, insumo_id, tipo)]
    total_entradas = sum(f["cantidad"] for f in filas if f["tipo"] == "Entrada")
    total_salidas = sum(f["cantidad"] for f in filas if f["tipo"] == "Salida")
    return {
        "movimientos": filas, "total_entradas": total_entradas, "total_salidas": total_salidas,
        "total_movimientos": len(filas),
    }


def informe_control_especial(fecha_desde: str = None, fecha_hasta: str = None):
    """
    Libro de control de medicamentos de control especial (los
    que se reportan al Fondo Nacional de Estupefacientes en
    Colombia) -- muestra cada movimiento de entrada y salida de
    estos medicamentos, con su lote, a qué paciente se le
    aplicó, qué profesional lo hizo, y el saldo que quedó, para
    tener la trazabilidad completa que exige la ley.
    """
    from database.database import consultar_todos

    sql = """
        SELECT m.*, i.nombre AS insumo_nombre, i.registro_invima, i.condicion_venta, i.unidad_medida,
               pr.nombre AS proveedor, pa.primer_nombre, pa.primer_apellido, pa.documento AS paciente_documento,
               pf.primer_nombre AS prof_nombre, pf.primer_apellido AS prof_apellido
        FROM inventario_movimientos m
        JOIN insumos i ON i.id = m.insumo_id
        LEFT JOIN proveedores pr ON pr.id = m.proveedor_id
        LEFT JOIN pacientes pa ON pa.id = m.paciente_id
        LEFT JOIN profesionales pf ON pf.id = m.profesional_id
        WHERE i.condicion_venta LIKE 'Control especial%'
    """
    parametros = []
    if fecha_desde:
        sql += " AND date(m.fecha) >= date(?)"
        parametros.append(fecha_desde)
    if fecha_hasta:
        sql += " AND date(m.fecha) <= date(?)"
        parametros.append(fecha_hasta)
    sql += " ORDER BY i.nombre, m.fecha"

    filas = [dict(f) for f in consultar_todos(sql, tuple(parametros))]
    for f in filas:
        f["paciente_nombre"] = f"{f.get('primer_nombre') or ''} {f.get('primer_apellido') or ''}".strip() or None
        f["profesional_nombre"] = f"{f.get('prof_nombre') or ''} {f.get('prof_apellido') or ''}".strip() or None

    return filas


def listar_insumos_control_especial():
    """Todos los medicamentos activos clasificados como de control especial -- para saber cuáles requieren este libro."""
    from database.database import consultar_todos
    filas = consultar_todos(
        "SELECT * FROM insumos WHERE activo=1 AND condicion_venta LIKE 'Control especial%' ORDER BY nombre"
    )
    return [dict(f) for f in filas]


def alertas_vencimiento(dias=90):
    """Lotes que vencen dentro de los próximos N días -- para gestionarlos antes de que se pierdan."""
    filas = [dict(f) for f in InventarioMovimientosRepository.lotes_por_vencer(dias)]
    for f in filas:
        f["dias_restantes"] = int(f["dias_restantes"]) if f["dias_restantes"] is not None else None
        f["vencido"] = f["dias_restantes"] is not None and f["dias_restantes"] < 0
    return filas


def resumen_dashboard():
    """Resumen ejecutivo para la pantalla principal de inventario."""
    existencias = informe_existencias()
    vencimientos = alertas_vencimiento(90)
    return {
        "total_insumos": existencias["total_insumos"],
        "total_valorizado": existencias["total_valorizado"],
        "insumos_stock_bajo": len(existencias["insumos_stock_bajo"]),
        "insumos_stock_exceso": len(existencias["insumos_stock_exceso"]),
        "lotes_por_vencer": len([v for v in vencimientos if not v["vencido"]]),
        "lotes_vencidos": len([v for v in vencimientos if v["vencido"]]),
        "total_convenios_vigentes": len([c for c in listar_convenios() if c["estado"] == "Vigente"]),
    }


# ==========================================
# CONVENIOS CON PROVEEDORES
# ==========================================

def listar_convenios():
    return [dict(c) for c in ConveniosRepository.listar_todos()]


def obtener_convenio(convenio_id: int):
    fila = ConveniosRepository.obtener(convenio_id)
    return dict(fila) if fila else None


def crear_convenio(proveedor_id, numero_convenio, tipo, fecha_inicio, fecha_fin, valor, condiciones, usuario_id) -> int:
    if not proveedor_id:
        raise ValueError("Debe indicar el proveedor.")
    if tipo not in TIPOS_CONVENIO:
        raise ValueError("Tipo de convenio no válido.")
    return ConveniosRepository.crear({
        "proveedor_id": proveedor_id, "numero_convenio": numero_convenio or "",
        "tipo": tipo, "fecha_inicio": fecha_inicio or None, "fecha_fin": fecha_fin or None,
        "valor": valor or None, "condiciones": condiciones or "", "usuario_creacion": usuario_id,
    })


def finalizar_convenio(convenio_id: int):
    ConveniosRepository.cambiar_estado(convenio_id, "Finalizado")
