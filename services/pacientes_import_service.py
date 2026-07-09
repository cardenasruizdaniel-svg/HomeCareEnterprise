"""
HomeCare Enterprise - Importacion masiva de pacientes desde Excel

Genera una plantilla .xlsx clara y facil de diligenciar (con
instrucciones, encabezados resaltados, un ejemplo, y listas
desplegables para los campos con valores fijos), y procesa el
archivo diligenciado al subirlo de vuelta.
"""

import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from core.texto_utils import normalizar
from repositories.catalogos_repository import DivipolaRepository
from services.pacientes_service import PacientesService

COLUMNAS = [
    ("tipo_documento", "Tipo Documento*", 14),
    ("documento", "Número Documento*", 18),
    ("primer_nombre", "Primer Nombre*", 18),
    ("segundo_nombre", "Segundo Nombre", 18),
    ("primer_apellido", "Primer Apellido*", 18),
    ("segundo_apellido", "Segundo Apellido", 18),
    ("fecha_nacimiento", "Fecha de Nacimiento* (AAAA-MM-DD)", 22),
    ("sexo", "Sexo*", 8),
    ("eps", "EPS", 20),
    ("regimen", "Régimen", 16),
    ("celular", "Celular", 14),
    ("telefono", "Teléfono fijo", 14),
    ("correo", "Correo electrónico", 24),
    ("direccion", "Dirección", 26),
    ("barrio", "Barrio", 18),
    ("municipio", "Municipio*", 18),
    ("departamento", "Departamento*", 18),
]

COLUMNAS_REQUERIDAS = ["tipo_documento", "documento", "primer_nombre", "primer_apellido", "fecha_nacimiento"]

AZUL = "0D6EFD"
GRIS_CLARO = "F4F6F9"


def plantilla_excel() -> bytes:

    wb = Workbook()

    # ---------------------------------------------------
    # HOJA 1: INSTRUCCIONES
    # ---------------------------------------------------

    hoja_instrucciones = wb.active
    hoja_instrucciones.title = "Instrucciones"

    instrucciones = [
        ("Cómo diligenciar esta plantilla", True, 14),
        ("", False, 11),
        ("1. Vaya a la pestaña 'Pacientes' (abajo).", False, 11),
        ("2. La primera fila ya tiene los nombres de columna — no la modifique ni la borre.", False, 11),
        ("3. La segunda fila es un EJEMPLO de cómo se ve un registro correcto — bórrela antes de subir el archivo.", False, 11),
        ("4. Escriba un paciente por fila, empezando en la fila 2.", False, 11),
        ("5. Los campos marcados con * son obligatorios.", False, 11),
        ("6. 'Tipo Documento' y 'Sexo' tienen una lista desplegable: haga clic en la celda y elija una opción.", False, 11),
        ("7. La fecha de nacimiento debe ir en formato AAAA-MM-DD, por ejemplo: 1980-05-20", False, 11),
        ("8. Guarde el archivo (no cambie el formato .xlsx) y súbalo en el sistema.", False, 11),
        ("", False, 11),
        ("El sistema le indicará, fila por fila, si algún dato quedó mal digitado, sin borrar lo que sí se pudo cargar.", False, 11),
    ]

    for fila, (texto, negrita, tamano) in enumerate(instrucciones, start=1):
        celda = hoja_instrucciones.cell(row=fila, column=1, value=texto)
        celda.font = Font(bold=negrita, size=tamano, color=AZUL if negrita else "000000")

    hoja_instrucciones.column_dimensions["A"].width = 100

    # ---------------------------------------------------
    # HOJA 2: PACIENTES (la que se diligencia)
    # ---------------------------------------------------

    hoja = wb.create_sheet("Pacientes")

    for col_idx, (clave, encabezado, ancho) in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=1, column=col_idx, value=encabezado)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=AZUL)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hoja.column_dimensions[chr(64 + col_idx)].width = ancho

    hoja.row_dimensions[1].height = 32
    hoja.freeze_panes = "A2"

    ejemplo = [
        "CC", "1094567890", "Carlos", "Andrés", "Ramírez", "Gómez",
        "1975-03-12", "M", "Sura", "Contributivo", "3009998877", "",
        "carlos@example.com", "Cll 10 # 5-20", "Centro", "Armenia", "Quindío",
    ]
    for col_idx, valor in enumerate(ejemplo, start=1):
        celda = hoja.cell(row=2, column=col_idx, value=valor)
        celda.font = Font(italic=True, color="888888")
        celda.fill = PatternFill("solid", fgColor=GRIS_CLARO)

    # Listas desplegables para evitar errores de digitación
    dv_tipo_doc = DataValidation(type="list", formula1='"RC,TI,CC,CE,PA,PEP,MSI,ASI"', allow_blank=True)
    dv_tipo_doc.error = "Elija un tipo de documento válido de la lista."
    hoja.add_data_validation(dv_tipo_doc)
    dv_tipo_doc.add("A2:A1000")

    dv_sexo = DataValidation(type="list", formula1='"M,F"', allow_blank=True)
    hoja.add_data_validation(dv_sexo)
    dv_sexo.add("H2:H1000")

    dv_regimen = DataValidation(type="list", formula1='"Contributivo,Subsidiado,Especial,Particular"', allow_blank=True)
    hoja.add_data_validation(dv_regimen)
    dv_regimen.add("J2:J1000")

    salida = io.BytesIO()
    wb.save(salida)
    return salida.getvalue()


def _leer_filas_excel(contenido_bytes: bytes):
    wb = load_workbook(io.BytesIO(contenido_bytes), data_only=True)

    if "Pacientes" in wb.sheetnames:
        hoja = wb["Pacientes"]
    else:
        hoja = wb.worksheets[0]

    filas = list(hoja.iter_rows(values_only=True))

    if not filas:
        raise ValueError("El archivo está vacío.")

    encabezados_archivo = [str(c).strip() if c else "" for c in filas[0]]
    mapa_encabezado_a_clave = {encabezado: clave for clave, encabezado, _ in COLUMNAS}
    claves_columnas = [mapa_encabezado_a_clave.get(enc, "") for enc in encabezados_archivo]

    if not any(claves_columnas):
        raise ValueError(
            "No se reconocen los encabezados del archivo. Use la plantilla descargada desde el sistema, "
            "sin modificar los nombres de columna de la primera fila."
        )

    registros = []
    for fila in filas[1:]:
        if fila is None or all(c is None or str(c).strip() == "" for c in fila):
            continue
        registro = {}
        for clave, valor in zip(claves_columnas, fila):
            if not clave:
                continue
            registro[clave] = "" if valor is None else str(valor).strip()
        registros.append(registro)

    return registros


def importar_pacientes_excel(contenido_bytes: bytes, usuario_id=None) -> dict:

    registros = _leer_filas_excel(contenido_bytes)

    exitosos = 0
    errores = []

    for numero_fila, datos in enumerate(registros, start=2):

        try:
            for campo in COLUMNAS_REQUERIDAS:
                if not datos.get(campo):
                    raise ValueError(f"falta el campo obligatorio '{campo}'")

            if datos.get("fecha_nacimiento"):
                datos["fecha_nacimiento"] = datos["fecha_nacimiento"][:10]

            if datos.get("municipio") and not datos.get("codigo_municipio_divipola"):
                coincidencias = [
                    dict(c) for c in DivipolaRepository.buscar_municipios(datos["municipio"], limite=10)
                ]

                if datos.get("departamento") and len(coincidencias) > 1:
                    depto_normalizado = normalizar(datos["departamento"])
                    coincidencias_filtradas = [
                        c for c in coincidencias
                        if depto_normalizado in normalizar(c["nombre_departamento"])
                    ]
                    if coincidencias_filtradas:
                        coincidencias = coincidencias_filtradas

                if coincidencias:
                    datos["codigo_municipio_divipola"] = coincidencias[0]["codigo_municipio"]

            PacientesService.guardar(datos)
            exitosos += 1

        except Exception as error:
            errores.append({"fila": numero_fila, "documento": datos.get("documento", ""), "error": str(error)})

    return {"exitosos": exitosos, "errores": errores, "total": exitosos + len(errores)}
